import time
import pandas as pd
import numpy as np
import sys
import datetime
import math
import xlrd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

class Logger(object):
    def __init__(self,filename='default.log',stream=sys.stdout):
        self.terminal = stream
        self.log = open(filename,'w')

    def write(self,message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass
sys.stdout=Logger('default.log',sys.stdout)
sys.stderr=Logger('error.log',sys.stderr)

#xtcsb=sys.argv[1]
xtcsb="G:\\32_DATA\\data_for_kaoqing\\20230421-20230421\\input\\系统参数表20230421-20230421.xlsx"
#xtcsb="C:\\Users\\Administrator\\Desktop\\kaoqing\\XBB\\人脸钉钉考勤\\data\\20230413-20230413\\input\\系统参数表20230413-20230413.xlsx"
print('------------------系统参数表路径-----------------------')
print(xtcsb)
print('------------------系统参数表路径-----------------------')

#读取系统参数表，转换为字典
dataWJ=pd.read_excel(xtcsb,usecols=['文件'])
data_array=np.array(dataWJ.stack())
list_WJ=data_array.tolist()
print('------------------参数表中文件-----------------------')
print(list_WJ)
print('------------------参数表中文件-----------------------')

dataLJ=pd.read_excel(xtcsb,usecols=['路径'])
data_array=np.array(dataLJ.stack())
list_LJ=data_array.tolist()
print('------------------参数表中路径-----------------------')
print(list_LJ)
print('------------------参数表中路径-----------------------')

dict_XTCSB=dict(zip(list_WJ,list_LJ))
print('------------文件及路径对应字典-------------')
print(dict_XTCSB)
print('------------文件及路径对应字典-------------')
#读取系统参数表，转换为字典

lj_kqcsb=dict_XTCSB['考勤参数表']
lj_kqyssj=dict_XTCSB['考勤原始数据']
lj_ddqdbb=dict_XTCSB['钉钉签到报表']
lj_zjjglj1=dict_XTCSB['中间结果路径1']
if not os.path.exists(lj_zjjglj1):
    os.makedirs(lj_zjjglj1)
lj_zjjglj2=dict_XTCSB['中间结果路径2']
if not os.path.exists(lj_zjjglj2):
    os.makedirs(lj_zjjglj2)
lj_zzjglj=dict_XTCSB['最终结果路径']
if not os.path.exists(lj_zzjglj):
    os.makedirs(lj_zzjglj)


print('##################################################################################################################')
print('#处理原始数据，每个人输出一个excel表格                                                                                 #')
print('##################################################################################################################')

#pandas读取人员名单
dataxm=pd.read_excel(lj_kqcsb,sheet_name='人员名单',usecols=['姓名'])
print(dataxm)
data_array=np.array(dataxm.stack())
list_XM=data_array.tolist()
print('------------人员名单-------------')
print(list_XM)
print('------------人员名单-------------')
#pandas读取人员名单

#pandas读取考勤日期加一天，并转化为列表
DF_KaoQinCanShu=pd.read_excel(lj_kqcsb,sheet_name='考勤日期')
list= DF_KaoQinCanShu.values[0]
startDay_KaoQin=list[0]
endDay_KaoQin=list[1]
endDay_addOneDay = (datetime.datetime.strptime(endDay_KaoQin,'%Y-%m-%d')+datetime.timedelta(days=1)).strftime('%Y-%m-%d')
list_RQ=[x.strftime('%Y-%m-%d') for x in pd.date_range(startDay_KaoQin,endDay_addOneDay)]
#pandas读取考勤日期加一天，并转化为列表

#考勤原始数据,一次性读取多个页签数据到一个dataframe中
df=pd.ExcelFile(lj_kqyssj)
df.sheet_names
data=pd.concat([pd.read_excel(df,sheet) for sheet in df.sheet_names])
print(data)
count_H=data.shape[0] #行数
count_L=data.shape[1] #列数

print("------------------------按照姓名分组输出数据-开始----------------------------------")
for xm in list_XM:
    data1=data[data['姓名']==xm]
    writer= pd.ExcelWriter(lj_zjjglj1+'/'+xm+'.xlsx')
    data2 = data1.sort_values(by='时间', ascending=True)
    j=0
    while(j <(len(list_RQ)-1)):
        data3=data2[(data2['时间']>list_RQ[j]) & (data2['时间']<list_RQ[j+1])]
        print(data3)
        data3.to_excel(writer,sheet_name=list_RQ[j],index=False)
        j=j+1
    writer.save()
    writer.close()
print("------------------------按照姓名分组输出数据-完成----------------------------------")



print('##################################################################################################################')
print('#                  休眠10秒钟，等待每人单独人脸数据写入文件                                                             #')
print('##################################################################################################################')
time.sleep(10)



print('##################################################################################################################')
print('#处理人脸数据，生成早到、迟到、早退、旷工以及待确认的一些数据                                                               #')
print('##################################################################################################################')
#pandas读取人员名单
dataxm=pd.read_excel(lj_kqcsb,sheet_name='人员名单',usecols=['姓名'])
print(dataxm)
data_array=np.array(dataxm.stack())
list_RYMD=data_array.tolist()
print('--------------------人员名单------------------------')
print(list_RYMD)
print('--------------------人员名单------------------------')
#pandas读取人员名单

#pandas读取考勤日期,并转化为列表
DF_KaoQinCanShu=pd.read_excel(lj_kqcsb,sheet_name='考勤日期')
list= DF_KaoQinCanShu.values[0]
startDay_KaoQin=list[0]
endDay_KaoQin=list[1]
list_SBRQ=[x.strftime('%Y-%m-%d') for x in pd.date_range(startDay_KaoQin,endDay_KaoQin)]
#pandas读取考勤日期,并转化为列表

print('--------------------上班日期------------------------')
print(list_SBRQ)
print('--------------------上班日期------------------------')
#pandas读取考勤日期


#pandas读取休假日期,并转化为列表
#休假区间
dataxj=pd.read_excel(lj_kqcsb,sheet_name='休假日期')
print('--------------------休假日期分段区间------------------------')
print(dataxj)
print('--------------------休假日期分段区间------------------------')
count_of_raw = dataxj.shape[0]
list_XJRQ=[]
l=0
while(l<count_of_raw):
    raw = dataxj.values[l]
    print(raw)
    startDay=raw[0]
    endDay=raw[1]
    days=[x.strftime('%Y-%m-%d') for x in pd.date_range(startDay,endDay)]
    print(days)
    list_XJRQ=list_XJRQ+days
    print(list_XJRQ)
    l=l+1
print('--------------------休假日期------------------------')
print(list_XJRQ)
print('--------------------休假日期------------------------')
#pandas读取休假日期,并转化为列表

#休假名单
DF_XJMDQJ=pd.read_excel(lj_kqcsb,sheet_name='休假名单')
print('--------------------休假名单区间------------------------')
print(DF_XJMDQJ)
print('--------------------休假名单区间------------------------')
#休假名单
count_of_raw = DF_XJMDQJ.shape[0]
list_qj_to_everyDay=[]
l=0
while(l<count_of_raw):
    raw = DF_XJMDQJ.values[l]
    print(raw)
    xm = raw[0]
    startDay=raw[1]
    endDay=raw[2]
    days=[x.strftime('%Y-%m-%d') for x in pd.date_range(startDay,endDay)]
    print(days)
    j=0
    len_of_days=len(days)
    while(j<len_of_days):
        dic_holiday = {}
        dic_holiday['姓名']=xm
        dic_holiday['休假日期']=days[j]
        list_qj_to_everyDay.append(dic_holiday)
        j=j+1
    l=l+1
print(list_qj_to_everyDay)
dataxjmd=pd.DataFrame(list_qj_to_everyDay)
print('--------------------休假名单------------------------')
print(dataxjmd)
print('--------------------休假名单------------------------')
#休假名单

#出差名单
DF_CCMDQJ=pd.read_excel(lj_kqcsb,sheet_name='出差名单')
print('--------------------出差名单区间------------------------')
print(DF_CCMDQJ)
print('--------------------出差名单区间------------------------')
count_of_raw = DF_CCMDQJ.shape[0]
list_qj_to_everyDay=[]
l=0
while(l<count_of_raw):
    raw = DF_CCMDQJ.values[l]
    print(raw)
    xm = raw[0]
    startDay=raw[1]
    endDay=raw[2]
    days=[x.strftime('%Y-%m-%d') for x in pd.date_range(startDay,endDay)]
    print(days)
    j=0
    len_of_days=len(days)
    while(j<len_of_days):
        dic_business = {}
        dic_business['姓名']=xm
        dic_business['出差日期']=days[j]
        list_qj_to_everyDay.append(dic_business)
        j=j+1
    l=l+1
print(list_qj_to_everyDay)
dataccmd=pd.DataFrame(list_qj_to_everyDay)
print('--------------------出差名单------------------------')
print(dataccmd)
print('--------------------出差名单------------------------')
#出差名单

# #营业厅人员名单，只统计早到
# datayyt=pd.read_excel(lj_kqcsb,sheet_name='营业厅',usecols=['姓名'])
# data_array=np.array(datayyt.stack())
# list_YYT=data_array.tolist()
# print('--------------------营业厅人员名单------------------------')
# print(list_YYT)
# print('--------------------营业厅人员名单------------------------')
# #营业厅人员名单，只统计早到

#只统计早到名单，只统计早到情况
DF_ZTJZDMDQJ=pd.read_excel(lj_kqcsb,sheet_name='只统计早到名单')
print('--------------------只统计早到名单区间------------------------')
print(DF_ZTJZDMDQJ)
print('--------------------只统计早到名单区间------------------------')
#
count_of_raw = DF_ZTJZDMDQJ.shape[0]
list_qj_to_everyDay=[]
l=0
while(l<count_of_raw):
    raw = DF_ZTJZDMDQJ.values[l]
    print(raw)
    xm = raw[1]
    startDay=raw[2]
    endDay=raw[3]
    days=[x.strftime('%Y-%m-%d') for x in pd.date_range(startDay,endDay)]
    print(days)
    j=0
    len_of_days=len(days)
    while(j<len_of_days):
        dic_holiday = {}
        dic_holiday['姓名']=xm
        dic_holiday['只统计早到日期']=days[j]
        list_qj_to_everyDay.append(dic_holiday)
        j=j+1
    l=l+1
print(list_qj_to_everyDay)
dataztjzdmd=pd.DataFrame(list_qj_to_everyDay)
print('--------------------只统计早到名单------------------------')
print(dataztjzdmd)
print('--------------------只统计早到名单------------------------')
#只统计早到名单，只统计早到情况

#ZH：记录当前人员是否在行里，0-不在，1-在行，初始化为0
ZH=0

#ZD：记录出勤是否早到，-1-未早到，0-早到待确认，1-早到，初始化为-1
ZD='未早到'

#ZC：记录出勤是否正常，-1-未正常，0-正常待确认，1-正常，初始化为-1
ZC='未正常'

#CD_SW：记录上午是否迟到，-1-迟到，0-未迟到待确认，1-未迟到，初始化为-1
CD_SW='迟到'

#CD_XW：记录下午是否迟到，-1-迟到，0-未迟到待确认，1-未迟到，初始化为-1
CD_XW='迟到'

#ZT_SW：记录上午是否早退，-1-早退，0-未早退待确认，1-未早退，初始化为-1
ZT_SW='早退'

#ZT_XW：记录下午是否早退，-1-早退，0-未早退待确认，1-未早退，初始化为-1
ZT_XW='早退'

#KG：记录是否旷工，0-旷工，1-未旷工待确认，2-未旷工，初始化为0
KG=0

#KG_SW：记录上午是否旷工，-1-旷工，0-未旷工待确认，1-未旷工，初始化为-1
KG_SW='旷工'

#KG_XW：记录下午是否旷工，-1-旷工，0-未旷工待确认，1-未旷工，初始化为-1
KG_XW='旷工'

#Tc_ZD：记录早到后出去的时间，初始化为00:00:00
Tc_ZD='00:00:00'

#Tr_ZD：记录早到出去后的第一个入时间，初始化为00:00:00
Tr_ZD='00:00:00'

#Tc_ZC：记录正常后出去的时间，初始化为00:00:00
Tc_ZC='00:00:00'

#Tr_ZC：记录正常出去后的第一个入时间，初始化为00:00:00
Tr_ZC='00:00:00'

#Tc_SB：记录上班后出去的时间，初始化为00:00:00
Tc_SB='00:00:00'

#Tc_ZTSW：记录上午早退的时间，初始化为空
Tc_ZTSW=''

#Tc_ZTXW：记录下午早退的时间，初始化为空
Tc_ZTXW=''

#Tr_SB：记录上班出去后的第一个入时间，初始化为00:00:00
Tr_SB='00:00:00'

#Tc：当前遍历的出时间，初始化为00:00:00
Tc='00:00:00'

#Tr：当前遍历的入时间，初始化为00:00:00
Tr='00:00:00'

#T：当前遍历的时间，初始化为00:00:00
T='00:00:00'

#List_WQ：记录需要外勤的时间段，每个时间段用一个字典表示{start:,end:}
List_WQ=[]

#考勤结果统计列表
list_KQTJ=[]

#统计每个人的早到次数
count_ZD=0

#早到统计列表
list_ZDTJ=[]

#缺勤统计
list_QQTJ=[]

#记录是否缺勤，1-代表缺勤，0-未缺勤，初始化为1
QQ=1

#记录上午迟到情况下9:30之前最后出的时间,打卡记录异常没有入则记录出的时间，空代表未迟到，初始化为空
Tc_CDSW=''

#记录上午迟到情况下第一次入（出）的时间,打卡记录异常没有入则记录出的时间，空代表未迟到，初始化为空
Tr_CDSW=''

#记录下午迟到情况下第一次入（出）的时间,打卡记录异常没有入则记录出的时间，空代表未迟到，初始化为空
Tr_CDXW=''

#出入超时时间段
list_CSSJD=[]

print('##################################################################################')
print('#                            遍历每个人每一条人脸数据                                 #')
print('##################################################################################')
for rymd in list_RYMD:
    print("\n-----------------------遍历人员：",rymd,'--------------------------------------')
    count_ZD=0 #每个人的早到次数，初始化为0
    for sbrq in list_SBRQ:
        # sfyc:记录当前入紧挨着的是否有出，0-紧挨着没有出，1-紧挨着有出，初始化为0
        sfyc = 0
        ZH = 0
        ZD = '未早到'
        ZC = '未正常'
        CD_SW = '迟到'
        CD_XW = '迟到'
        ZT_SW = '早退'
        ZT_XW = '早退'
        KG = 0
        KG_SW = '旷工'
        KG_XW = '旷工'
        QQ=1
        sfytj_zdcs=0 #是否已经统计早到超时，0代表未统计，1代表已统计

        Tc_ZD = '00:00:00'
        Tr_ZD = '00:00:00'
        Tc_ZC = '00:00:00'
        Tr_ZC = '00:00:00'
        Tc_SB = '00:00:00'
        Tr_SB = '00:00:00'
        Tc_ZTSW=''
        Tc_ZTXW = ''
        Tc = '00:00:00'
        Tr = '00:00:00'
        Tc_ZD_CS='00:00:00'
        Tc_CDSW=''
        Tr_CDSW=''
        Tr_CDXW=''

        print('遍历日期：',sbrq)
        data=pd.read_excel(lj_zjjglj1+'/'+rymd+'.xlsx',sheet_name=sbrq)
        print(rymd,sbrq,data.info())
        i = 0
        # 数据行数
        count_H = data.shape[0]
        while(i<count_H):
            print('------第',i+1,'行人脸记录--------')
            list=data.values[i]
            print(list)
            # 出入标志
            crbz=list[6]
            #出入时间
            t=list[4]
            Y=t[0:4]#年份
            M=t[5:7]#月份
            D=t[8:10]#日期
            h=t[11:13]#小时
            m=t[14:16]#分钟
            s=t[17:19]#秒钟
            T=t[11:19]#时分秒
            print('出入时间：',T)
            if("入口" in crbz):
                ZH=1
                QQ=0
                if (sfyc == 1):
                    Tc_Y = Tc[0:4]  # 年份
                    Tc_M = Tc[5:7]  # 月份
                    Tc_D = Tc[8:10]  # 日期
                    Tc_h = Tc[11:13]  # 小时
                    Tc_m = Tc[14:16]  # 分钟
                    Tc_s = Tc[17:19]  # 秒钟

                    if(T<='09:01:00'):
                        ZD='早到'

                    #早上9:00前签到后离行超过10分钟
                    if(t>Tc_ZD_CS and ZD=='待确认' and sfytj_zdcs==0):
                        # if ((rymd not in list_YYT) and (dataxjmd[(dataxjmd['姓名'] == rymd) & (dataxjmd['休假日期'] ==
                        #                                                                      sbrq)].empty == True)):
                        if((dataztjzdmd[(dataztjzdmd['姓名'] == rymd) & (dataztjzdmd['只统计早到日期'] ==
                                                                                            sbrq)].empty==True) and (dataxjmd[(dataxjmd['姓名'] == rymd) & (dataxjmd['休假日期'] ==
                                                                                            sbrq)].empty==True)):
                            dic_CSSJD = {}
                            dic_CSSJD["姓名"] = rymd
                            dic_CSSJD["日期"] = sbrq
                            dic_CSSJD["出去"] = Tc_ZD[11:19]
                            dic_CSSJD["进来"] = t[11:19]
                            start_time = datetime.datetime.strptime(Tc_ZD, '%Y-%m-%d %H:%M:%S')
                            end_time=datetime.datetime.strptime(t,'%Y-%m-%d %H:%M:%S')
                            dic_CSSJD["时间差(分钟)"]=math.floor(((end_time-start_time).seconds)/60)
                            print("早到出入超时啦，出：",Tc_ZD,'入：',t)
                            dic_CSSJD["早到出入超时"] = 1
                            list_CSSJD.append(dic_CSSJD)
                            Tr_ZD=t
                            sfytj_zdcs=1

                    else:
                        Tc_CS = (datetime.datetime(int(Tc_Y), int(Tc_M), int(Tc_D), int(Tc_h), int(Tc_m),
                                                  int(Tc_s)) + datetime.timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')  # 超时时间等于上次出的时间加上30分钟

                        print("上次出的时间是:", Tc)
                        print("出去的超时时间为:", Tc_CS)
                        if(t>Tc_CS):
                            print("此次入距离上次出超时啦,超时时段",Tc,t)
                            # if ((rymd not in list_YYT) and (dataxjmd[(dataxjmd['姓名'] == rymd) & (dataxjmd['休假日期'] ==
                            #                                                                      sbrq)].empty == True))
                            if((dataztjzdmd[(dataztjzdmd['姓名'] == rymd) & (dataztjzdmd['只统计早到日期'] ==
                                                                                            sbrq)].empty==True) and (dataxjmd[(dataxjmd['姓名'] == rymd) & (dataxjmd['休假日期'] ==
                                                                                                sbrq)].empty==True)):
                                dic_CSSJD={}
                                dic_CSSJD["姓名"]=rymd
                                dic_CSSJD["日期"]=sbrq
                                dic_CSSJD["出去"]=Tc[11:19]
                                dic_CSSJD["进来"]=t[11:19]
                                start_time = datetime.datetime.strptime(Tc, '%Y-%m-%d %H:%M:%S')
                                end_time = datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S')
                                dic_CSSJD["时间差(分钟)"]=math.floor(((end_time-start_time).seconds)/60)
                                if(Tc_ZC != '00:00:00'):
                                    print("正常出入超时")
                                    dic_CSSJD["正常出入超时"] = 1
                                    list_CSSJD.append(dic_CSSJD)
                                elif(Tc_SB != '00:00:00'):
                                    print("上班出入超时")
                                    dic_CSSJD["上班出入超时"] = 1
                                    list_CSSJD.append(dic_CSSJD)
                sfyc=0 #修改已出标志为未出
                if(T<'07:00:00'):
                    i=i
                elif(T<='09:01:00'):
                    ZD='早到'
                    ZC='正常'
                    CD_SW='未迟到'
                    CD_XW='未迟到'
                    ZT_SW='未早退'
                    ZT_XW='未早退'
                    KG_SW='未旷工'
                    KG_XW='未旷工'
                    Tc_ZTSW = ''
                    Tc_ZTXW = ''
                    sfyc=0
                elif(('09:01:00'<T) & (T<='09:31:00')):
                    ZC='正常'
                    CD_SW='未迟到'
                    CD_XW='未迟到'
                    ZT_SW='未早退'
                    ZT_XW='未早退'
                    KG_SW='未旷工'
                    KG_XW='未旷工'
                    Tc_ZTSW = ''
                    Tc_ZTXW = ''
                    sfyc=0
                elif(('09:31:00'<T) & (T<='13:00:00')):
                    CD_XW='未迟到'
                    ZT_SW='未早退'
                    ZT_XW='未早退'
                    KG_SW='未旷工'
                    KG_XW='未旷工'
                    Tc_ZTSW = ''
                    Tc_ZTXW = ''
                    sfyc=0
                    # if ((CD_SW == '待确认') & (Tr_CDSW == '')):
                    #     Tr_CDSW = t
                    if (Tr_CDSW == ''):
                        Tr_CDSW = t
                #20220519，根据人力的规则进行调整，15:40后才算迟到
                elif(('13:00:00'<T) & (T<='15:41:00')):
                    CD_XW='未迟到'
                    ZT_XW='未早退'
                    KG_XW='未旷工'
                    Tc_ZTXW = ''
                    sfyc=0
                elif(('15:41:00'<T) & (T<='18:30:00')):
                    ZT_XW='未早退'
                    KG_XW='未旷工'
                    Tc_ZTXW = ''
                    sfyc=0
                    if(CD_XW !='未迟到'):
                        CD_XW='待确认'
                    if((CD_XW=='待确认') & (Tr_CDXW=='')):
                        Tr_CDXW=t
            else:
                ZH=0
                sfyc=1
                Tc=t#记录出去的时间
                if (T < '07:00:00'):
                    i=i
                elif (T <= '09:00:00'):
                    ZD = '待确认'
                    ZC = '待确认'
                    CD_SW = '待确认'
                    CD_XW='待确认'
                    ZT_SW='待确认'
                    ZT_XW='待确认'
                    KG_XW='待确认'
                    Tc_ZD=t
                    Tc_CDSW=t
                    Tc_ZD_CS=(datetime.datetime(int(Y), int(M), int(D), int(h), int(m),
                                              int(s)) + datetime.timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S')  # 早上9:00前签到后离行超过10分钟
                    Tc_ZC='00:00:00'
                    Tc_SB='00:00:00'
                    Tc_ZTSW=t
                    Tc_ZTXW=t
                elif (('09:00:00' < T) & (T <= '09:30:00')):
                    ZC = '待确认'
                    CD_SW = '待确认'
                    CD_XW='待确认'
                    ZT_SW='待确认'
                    ZT_XW='待确认'
                    KG_XW='待确认'
                    Tc_ZC=t
                    Tc_CDSW=t
                    # Tc_ZD='00:00:00'
                    Tc_SB='00:00:00'
                    Tc_ZTSW = t
                    Tc_ZTXW = t
                elif (('09:30:00' < T) & (T <= '12:55:00')):
                    Tc_SB=t
                    Tc_ZTSW = t
                    Tc_ZTXW = t
                    CD_XW='待确认'
                    ZT_SW='待确认'
                    ZT_XW='待确认'
                    KG_XW = '待确认'
                    # Tc_ZD='00:00:00'
                    Tc_ZC='00:00:00'
                    # if ((CD_SW == '待确认') & (Tr_CDSW == '')):
                    #     Tr_CDSW = t
                elif (('12:55:00' < T) & (T <= '13:00:00')):
                    Tc_SB=t
                    CD_XW='待确认'
                    ZT_SW='未早退'
                    ZT_XW='待确认'
                    KG_XW = '待确认'
                    Tc_ZTXW = t
                    # Tc_ZD='00:00:00'
                    Tc_ZC='00:00:00'
                    # if ((CD_SW == '待确认') & (Tr_CDSW == '')):
                    #     Tr_CDSW = t
                elif (('13:00:00' < T) & (T <= '15:20:00')):
                    CD_XW='待确认'
                    ZT_XW='待确认'
                    KG_XW = '待确认'
                    Tc_ZTXW = t
                    # Tc_ZD = '00:00:00'
                    Tc_ZC = '00:00:00'
                    Tc_SB = '00:00:00'
                elif (('15:20:00' < T) & (T <= '15:40:00')):
                    pass
                elif(('15:40:00' < T) & (T <= '18:25:00')):
                    Tc_SB=t
                    ZT_XW = '待确认'
                    Tc_ZTXW = t
                    KG_XW = '未旷工'
                    # Tc_ZD='00:00:00'
                    Tc_ZC='00:00:00'
                    # if ((CD_XW == '待确认') & (Tr_CDXW == '')):
                    #     Tr_CDXW = t
                elif(('18:25:00' < T) & (T <= '18:30:00')):
                    Tc_SB=t
                    ZT_XW = '未早退'
                    KG_XW = '未旷工'
                    # Tc_ZD='00:00:00'
                    Tc_ZC='00:00:00'
                    # if ((CD_XW == '待确认') & (Tr_CDXW == '')):
                    #     Tr_CDXW = t
                else:
                    # Tc_ZD = '00:00:00'
                    Tc_ZC = '00:00:00'
                    Tc_SB = '00:00:00'
            i=i+1

        # 早上9:00前签到离行后未入
        if (ZD == '待确认' and sfytj_zdcs == 0):
            # if (rymd not in list_YYT):
            if (dataztjzdmd[(dataztjzdmd['姓名'] == rymd) & (dataztjzdmd['只统计早到日期'] == sbrq)].empty==True):
                dic_CSSJD = {}
                dic_CSSJD["姓名"] = rymd
                dic_CSSJD["日期"] = sbrq
                dic_CSSJD["出去"] = Tc_ZD[11:19]
                dic_CSSJD["进来"] = ""
                dic_CSSJD["时间差(分钟)"] = ''
                print("早到出入超时")
                dic_CSSJD["早到出入超时"] = 1
                list_CSSJD.append(dic_CSSJD)
                sfytj_zdcs = 1

        # 考勤结果统计字典
        if(sbrq not in list_XJRQ):
            dic_KQTJ = {}
            print(rymd,sbrq,"是否早到：",ZD)
            dic_KQTJ["姓名"]=rymd
            dic_KQTJ["日期"]=sbrq
            dataxjmd_mm = dataxjmd[(dataxjmd['姓名'] == rymd) & (dataxjmd['休假日期'] == sbrq)]
            dataccmd_mm = dataccmd[(dataccmd['姓名'] == rymd) & (dataccmd['出差日期'] == sbrq)]
            #休假情况，都不统计
            if (dataxjmd_mm.empty == False):
                #print(dataxjmd_mm)
                dic_KQTJ["休假"]="休假"
            # 出差情况，都不统计
            elif (dataccmd_mm.empty == False):
                dic_KQTJ["休假"] = "出差"
            # 未休假、未出差
            else:
                dic_KQTJ["休假"] = "未休"
                dic_KQTJ["早到"]=ZD
                #营业厅人员只统计是否早到
                # if (rymd in list_YYT):
                if (dataztjzdmd[(dataztjzdmd['姓名'] == rymd) & (dataztjzdmd['只统计早到日期'] == sbrq)].empty==False):
                    if(ZD=='早到' or ZD=='待确认'):
                        dic_KQTJ["早到"] = '早到'
                else:
                    dic_KQTJ["早到出"]=Tc_ZD[11:19]
                    dic_KQTJ["早到入"]=Tr_ZD[11:19]
                    if(ZD!='待确认'):
                        dic_KQTJ["早到出"] = ''
                        dic_KQTJ["早到入"] = ''
                    dic_KQTJ["正常"]=ZC
                    dic_KQTJ["上午迟到"] = CD_SW
                    if(CD_SW!='待确认'):
                       dic_KQTJ["上午迟到出"] = ''
                    else:
                       dic_KQTJ["上午迟到出"] = Tc_CDSW[11:19]
                    dic_KQTJ["上午迟到入"] = Tr_CDSW[11:19]
                    if(CD_SW=='未迟到'):
                        dic_KQTJ["上午迟到入"] = ''
                    dic_KQTJ["下午迟到"] = CD_XW
                    dic_KQTJ["下午迟到入"] = Tr_CDXW[11:19]
                    dic_KQTJ["上午早退"] = ZT_SW
                    dic_KQTJ["上午早退出"] = Tc_ZTSW[11:19]
                    dic_KQTJ["下午早退"] = ZT_XW
                    dic_KQTJ["下午早退出"] = Tc_ZTXW[11:19]
                    dic_KQTJ["上午旷工"] = KG_SW
                    if(KG_SW=='旷工'):
                        dic_KQTJ["上午迟到"] = ''
                        dic_KQTJ["上午早退"] = ''
                        dic_KQTJ["早到"] = ''
                        dic_KQTJ["正常"] = ''
                        dic_KQTJ["上午早退出"] = ''
                    if(KG_XW=='旷工'):
                        dic_KQTJ["下午迟到"] = ''
                        dic_KQTJ["下午早退"] = ''
                        dic_KQTJ["下午早退出"] = ''
                    dic_KQTJ["下午旷工"] = KG_XW
                    # if(ZD=='待确认' or ZD=='早到'):
                    #     count_ZD=count_ZD+1
            list_KQTJ.append(dic_KQTJ)
            print(rymd, sbrq, "上午旷工：", KG_SW)
            print(rymd,sbrq,"正常：",ZC)
            print(rymd,sbrq,"上午迟到：",CD_SW)
            print(rymd, sbrq, "上午早退：", ZT_SW)
            print(rymd, sbrq, "下午旷工：", KG_XW)
            print(rymd,sbrq,"下午迟到：",CD_XW)
            print(rymd, sbrq, "下午早退：", ZT_XW)

            # 缺勤统计字典
            dic_QQTJ = {}
            dic_QQTJ["姓名"] = rymd
            dic_QQTJ["日期"]=sbrq
            dic_QQTJ["是否缺勤"] = QQ
            list_QQTJ.append(dic_QQTJ)

    # 早到统计字典
    # dic_ZDTJ = {}
    # dic_ZDTJ["姓名"]=rymd
    # dic_ZDTJ["早到次数"]=count_ZD
    # list_ZDTJ.append(dic_ZDTJ)

print(list_KQTJ)
data_KQTJ=pd.DataFrame(list_KQTJ)
print('---------中间考勤统计结果-只参考人脸数据----------')
print(data_KQTJ)
writer = pd.ExcelWriter(lj_zjjglj2+'/考勤统计结果.xlsx')
data_KQTJ.to_excel(writer,index=False)
writer.save()
writer.close()
print('---------中间考勤统计结果-只参考人脸数据----------')


print(list_QQTJ)
data_QQTJ=pd.DataFrame(list_QQTJ)
print('---------中间缺勤统计结果-只参考人脸数据----------')
print(data_QQTJ)
writer = pd.ExcelWriter(lj_zjjglj2+'/缺勤统计结果.xlsx')
data_QQTJ.to_excel(writer,index=False)
writer.save()
writer.close()
print('---------中间缺勤统计结果-只参考人脸数据----------')


print(list_CSSJD)
data_CSSJD=pd.DataFrame(list_CSSJD)
print('---------中间出入超时统计结果-只参考人脸数据----------')
print(data_CSSJD)
writer = pd.ExcelWriter(lj_zjjglj2+'/出入超时统计.xlsx')
data_CSSJD.to_excel(writer,index=False)
writer.save()
writer.close()
print('---------中间出入超时统计结果-只参考人脸数据----------')

print('################################################################################################################')
print('#                   休眠10秒钟，等待考勤中间结果写入到文件                                                            #')
print('################################################################################################################')
time.sleep(10)



print('################################################################################################################')
print('#                     将人脸统计数据与钉钉签到数据进行对比,与事假数据进行对比 ，处理一些待确认事项                           #')
print('################################################################################################################')
#人脸统计数据
dataFace=pd.read_excel(lj_zjjglj2+"/考勤统计结果.xlsx")
print("-------------------------------考勤统计中间结果------------------------------")
print(dataFace)
print("-------------------------------考勤统计中间结果------------------------------")

#出入超时数据
dataOvertime=pd.read_excel(lj_zjjglj2+"/出入超时统计.xlsx")
print("-------------------------------考出入超时统计中间结果------------------------------")
print(dataOvertime)
print("-------------------------------考出入超时统计中间结果------------------------------")

#钉钉签到数据
print("-------------------------------钉钉签到数据------------------------------")
#定义目录路径
path = lj_ddqdbb
#通过函数取出所有文件名称
files = os.listdir(path)
dataDingding=pd.DataFrame()
#循环拼接路径>读取所有文件
for i in range(0,len(files)):
    file_name=path+files[i]
    df1=pd.read_excel(file_name)
    frames=[dataDingding,df1]
    dataDingding=pd.concat(frames)
print(dataDingding)
print("-------------------------------钉钉签到数据------------------------------")

#事假数据
dataShijia=pd.read_excel(lj_kqcsb,sheet_name='事假名单')
print("-------------------------------事假名单数据------------------------------")
print(dataShijia)
print("-------------------------------事假名单数据------------------------------")

#数据行数,人脸数据
count_H = dataFace.shape[0]
i=0

#数据行数，出入超时数据
count_Overtime = dataOvertime.shape[0]
k=0

#钉钉数据确认后的考勤统计表
list_FaceDingding=[]

#出入超时钉钉签到情况
list_OvertimeDingding=[]

print('--------遍历出入超时中间统计结果----------')
while(k<count_Overtime):
    list = dataOvertime.values[k]
    xm=list[0]
    rq=list[1]
    offTime = list[2]
    entyTime=list[3]
    #zdcsbz=list[6] #早到出入超时标志

    #出入时间区间要么在上午上班时间区间，要么在下午上班时间区间
    if((offTime>'09:30:00' and entyTime<'13:00:00') or (offTime>'15:30:00' and entyTime<'18:30:00') ):
        dic_OvertimeDingding={}
        dic_OvertimeDingding["姓名"]=xm
        dic_OvertimeDingding["日期"] = rq
        dic_OvertimeDingding["出"]=offTime
        dic_OvertimeDingding["入"] = entyTime
        dic_OvertimeDingding["时间差（分钟）"] = list[4]
        dic_OvertimeDingding["外勤卡"] = ''
        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        #处理钉钉数据
        print('\n第',k+1,'出入超时记录',"姓名：", xm, "日期：", rq, "开始时间：", offTime, "结束时间：", entyTime,'\n开始寻找钉钉签到数据')
        ddjlData = dataDingding[(dataDingding['姓名'] == xm) & (dataDingding['日期'] == rq)]
        if (ddjlData.empty == False):
            print(ddjlData)
            count_ddjl = ddjlData.shape[0]
            j = 0
            while (j < count_ddjl):
                ddjl = ddjlData.values[j]
                dksj = ddjl[4] #钉钉打卡时间
                if (entyTime > dksj and offTime < dksj):
                    print("打卡时间：", dksj, "在出去时间：", offTime, "进来时间：", entyTime, "之间")
                    dic_OvertimeDingding["外勤卡"] = dksj
                    wqk=1
                    print('寻找到符合要求的钉钉签到数据')
                    break
                j = j + 1
        print('结束寻找钉钉签到数据')

        #处理事假数据
        print('开始寻找事假数据')
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            print("事假记录筛选")
            print(shijiajlData)
            print("事假记录筛选")
            count_shijiajl = shijiajlData.shape[0]
            print("一共筛选了：", count_shijiajl, "条事假记录")
            t = 0
            while (t < count_shijiajl):
                shijiajl = shijiajlData.values[t]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                print("事假开始时间:", startTime, "事假结束时间：", endTime)
                if (startTime <= offTime and endTime >= entyTime):
                    wqk = 1
                    print("命中休假数据，算作外勤卡")
                    dic_OvertimeDingding["外勤卡"] = '事假'
                    break
                t = t + 1
        print('结束寻找事假数据')
        if(wqk==0):
           list_OvertimeDingding.append(dic_OvertimeDingding) #出入超时缺钉钉卡的数据
           print('************没能寻找到符合要求的钉钉签到数据和事假数据***********')
    k=k+1

print('------------遍历人脸统计中间结果-----')
list_XYWQKSJD=[]
while(i<count_H):
    dic_FaceDingding={}
    list=dataFace.values[i]
    dic_FaceDingding["姓名"]=list[0]
    dic_FaceDingding["日期"]=list[1]
    dic_FaceDingding["休假"]=list[2]
    dic_FaceDingding["上午旷工"] = list[16]
    dic_FaceDingding["上午旷工卡"] = ''
    dic_FaceDingding["下午旷工"] = list[17]
    dic_FaceDingding["下午旷工卡"] = ''
    dic_FaceDingding["早到"]=list[3]
    dic_FaceDingding["早到出"] = list[4]
    dic_FaceDingding["早到入"] = list[5]
    dic_FaceDingding["早到卡"] = ''
    #dic_FaceDingding["正常"]=list[6]
    dic_FaceDingding["上午迟到"]=list[7]
    dic_FaceDingding["上午迟到出"] = list[8]
    dic_FaceDingding["上午迟到入"] = list[9]
    dic_FaceDingding["上午迟到卡"] =""
    dic_FaceDingding["下午迟到"] = list[10]
    dic_FaceDingding["下午迟到入"] = list[11]
    dic_FaceDingding["下午迟到卡"] = ''
    dic_FaceDingding["上午早退"] = list[12]
    dic_FaceDingding["上午早退出"] = list[13]
    dic_FaceDingding["上午早退卡"] = ""
    dic_FaceDingding["下午早退"] = list[14]
    dic_FaceDingding["下午早退出"] = list[15]
    dic_FaceDingding["下午早退卡"] = ""

    xm = list[0]
    rq = list[1]
    ZDC_SJ=str(list[4])#早到待确认出时间
    ZDR_SJ = str(list[5]) #早到待确认入时间
    SWZT_SJ=str(list[13]) #上午早退时间，钉钉签到时间在早退时间与13:00之间
    XWZT_SJ=str(list[15]) #下午早退时间，钉钉签到时间在早退时间与18:30之间
    SWCD=str(list[7]) #上午迟到情况
    SWZT=str(list[12]) #上午早退情况
    XWCD=str(list[10]) #下午迟到情况
    XWZT=str(list[14]) #下午早退情况
    SWCDC_SJ=str(list[8]) #上午迟到出时间
    SWCDR_SJ=str(list[9]) #上午迟到入时间
    XWCDR_SJ=str(list[11]) #下午迟到入时间
    SWKG=str(list[16]) #上午旷工情况
    XWKG=str(list[17]) #下午旷工情况

    print("上午早退时间",SWZT_SJ)
    print("下午早退时间",XWZT_SJ)

    print('第',i+1,'条中间结果数据')

    print('------------- 开始处理上午早到待确认数据，与钉钉匹配签到数据 ---------')
    if (ZDC_SJ != 'nan'):
        dic_WQKSJD = {}
        dic_WQKSJD["姓名"] = xm
        dic_WQKSJD["日期"] = rq
        dic_WQKSJD["开始"] = ZDC_SJ
        if(ZDR_SJ == 'nan'):
            dic_WQKSJD["结束"] = ''
            ZDR_SJ='18:30:00'
        dic_WQKSJD["结束"] = ZDR_SJ
        dic_WQKSJD["签到"] = ''
        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        ddjlData = dataDingding[(dataDingding['姓名'] == xm) & (dataDingding['日期'] == rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：", ZDC_SJ, "结束时间：", ZDR_SJ)
            print(ddjlData)
            count_ddjl = ddjlData.shape[0]
            j = 0
            while (j < count_ddjl):
                ddjl = ddjlData.values[j]
                dksj = ddjl[4]
                if (ZDR_SJ > dksj and ZDC_SJ < dksj):
                    print("打卡时间：", dksj, "在出去时间：", ZDC_SJ, "进来时间：", ZDR_SJ, "之间")
                    dic_WQKSJD["签到"] = dksj
                    wqk = 1
                    dic_FaceDingding["早到"] = "早到-钉钉"
                    dic_FaceDingding["早到卡"] = dksj
                    break
                j = j + 1
        if (wqk == 0):
            dic_FaceDingding["早到"] = "未早到-钉钉"
        list_XYWQKSJD.append(dic_WQKSJD)
    print('------------- 结束处理上午早到待确认数据，与钉钉匹配签到数据 ---------')


    print('\n------------- 开始处理上午早退待确认数据，与钉钉匹配签到数据，匹配事假数据 ---------')
    if(SWZT_SJ!='nan'):
        dic_WQKSJD={}
        dic_WQKSJD["姓名"]=xm
        dic_WQKSJD["日期"]=rq
        dic_WQKSJD["开始"]=SWZT_SJ
        dic_WQKSJD["结束"]='15:30:00'
        dic_WQKSJD["签到"]=''
        wqk=0  #外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        ddjlData=dataDingding[(dataDingding['姓名']==xm) & (dataDingding['日期']==rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：",SWZT_SJ, "结束时间：", "15:30:00")
            print(ddjlData)
            count_ddjl=ddjlData.shape[0]
            j=0
            while(j<count_ddjl):
                ddjl=ddjlData.values[j]
                dksj=ddjl[4]
                if('15:30:00'>dksj and SWZT_SJ<dksj):
                    print("打卡时间：",dksj,"在出去时间：",SWZT_SJ,"进来时间：","15:30:00","之间")
                    dic_WQKSJD["签到"]=dksj
                    wqk=1
                    dic_FaceDingding["上午早退"]="未早退-钉钉"
                    dic_FaceDingding["上午早退卡"] = dksj
                    break
                j=j+1
        if(wqk==0):
            dic_FaceDingding["上午早退"] = "早退-钉钉"
        list_XYWQKSJD.append(dic_WQKSJD)

        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            k = 0
            while (k < count_shijiajl):
                shijiajl = shijiajlData.values[k]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                if (startTime <= SWZT_SJ and endTime >= SWZT_SJ):
                    shijia = 1
                    dic_FaceDingding["上午早退"] = "未早退-事假"
                    break
                k = k + 1
    print('------------- 结束处理上午早退待确认数据，与钉钉匹配签到数据，匹配事假数据 ---------')


    print('\n------------- 开始处理下午早退待确认数据，与钉钉匹配签到数据，匹配事假数据 ---------')
    if(XWZT_SJ!='nan'):
        dic_WQKSJD={}
        dic_WQKSJD["姓名"]=xm
        dic_WQKSJD["日期"]=rq
        dic_WQKSJD["开始"]=XWZT_SJ
        dic_WQKSJD["结束"]='23:59:59'
        dic_WQKSJD["签到"] = ''
        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        XWYX_SJ='23:59:59'
        ddjlData=dataDingding[(dataDingding['姓名']==xm) & (dataDingding['日期']==rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：",XWZT_SJ, "结束时间：", XWYX_SJ)
            print(ddjlData)
            count_ddjl=ddjlData.shape[0]
            #下午没来，早退时间按上午早退时间算
            if(XWZT_SJ<'15:30:00'):
                XWZT_SJ='15:30:00'
                XWYX_SJ='18:30:00'
                dic_WQKSJD["结束"] = '23:59:59'
            j=0
            while(j<count_ddjl):
                ddjl=ddjlData.values[j]
                dksj=ddjl[4]
                if(XWYX_SJ>dksj and XWZT_SJ<dksj):
                    print("打卡时间：",dksj,"在出去时间：",XWZT_SJ,"进来时间：",XWYX_SJ,"之间")
                    dic_WQKSJD["签到"]=dksj
                    wqk=1
                    dic_FaceDingding["下午早退"] = "未早退-钉钉"
                    dic_FaceDingding["下午早退卡"] = dksj
                    break
                j=j+1
        if(wqk==0):
            dic_FaceDingding["下午早退"] = "早退-钉钉"
        list_XYWQKSJD.append(dic_WQKSJD)

        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            k = 0
            while (k < count_shijiajl):
                shijiajl = shijiajlData.values[k]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                if (startTime <= XWZT_SJ and endTime >= XWZT_SJ):
                    shijia = 1
                    dic_FaceDingding["下午早退"] = "未早退-事假"
                    break
                k = k + 1
    print('------------- 结束处理下午早退待确认数据，与钉钉匹配签到数据，匹配事假数据 ---------')


    print('\n------------- 开始处理上午迟到待确认数据，与钉钉匹配签到数据，与事假数据匹配-------------')
    if (SWCDC_SJ != 'nan'):
        dic_WQKSJD = {}
        dic_WQKSJD["姓名"] = xm
        dic_WQKSJD["日期"] = rq
        dic_WQKSJD["开始"] = SWCDC_SJ
        if(SWCDR_SJ == 'nan'):
            SWCDR_SJ='13:00:00'
        dic_WQKSJD["结束"] = SWCDR_SJ
        dic_WQKSJD["签到"] = ''
        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        ddjlData = dataDingding[(dataDingding['姓名'] == xm) & (dataDingding['日期'] == rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：", SWCDC_SJ, "结束时间：", SWCDR_SJ)
            print(ddjlData)
            count_ddjl = ddjlData.shape[0]
            j = 0
            while (j < count_ddjl):
                ddjl = ddjlData.values[j]
                dksj = ddjl[4]
                if (SWCDR_SJ > dksj and SWCDC_SJ < dksj):
                    print("打卡时间：", dksj, "在出去时间：", SWCDC_SJ ,"进来时间：", SWCDR_SJ, "之间")
                    dic_WQKSJD["签到"] = dksj
                    wqk = 1
                    dic_FaceDingding["上午迟到"] = "未迟到-钉钉"
                    #dic_FaceDingding["正常"] = "正常-钉钉"
                    dic_FaceDingding["上午迟到卡"] = dksj
                    break
                j = j + 1
        if (wqk == 0):
            dic_FaceDingding["上午迟到"] = "迟到-钉钉"
            #dic_FaceDingding["正常"] = "未正常-钉钉"
        list_XYWQKSJD.append(dic_WQKSJD)

        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            k = 0
            while (k < count_shijiajl):
                shijiajl = shijiajlData.values[k]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                if (startTime == '09:30:00' and endTime >= SWCDR_SJ):
                    shijia = 1
                    dic_FaceDingding["上午迟到"] = "未迟到-事假"
                    break
                k = k + 1
    print('------------- 结束处理上午迟到待确认数据，与钉钉匹配签到数据，与事假数据匹配-------------')


    print('\n---------------------- 开始处理上午迟到数据，与事假数据匹配-------------------------')
    if (SWCD == '迟到'):
        if (SWCDR_SJ == 'nan'):
            SWCDR_SJ = '13:00:00'
        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            k = 0
            while (k < count_shijiajl):
                shijiajl = shijiajlData.values[k]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                if (startTime == '09:30:00' and endTime >= SWCDR_SJ):
                    shijia = 1
                    dic_FaceDingding["上午迟到"] = "未迟到-事假"
                    break
                k = k + 1
    print('---------------------- 结束处理上午迟到数据，与事假数据匹配-------------------------')


    print('\n---------------------- 开始处理下午迟到待确认数据，与钉钉匹配签到数据和事假数据-------------------------')
    if (XWCD=='待确认'):
        dic_WQKSJD = {}
        dic_WQKSJD["姓名"] = xm
        dic_WQKSJD["日期"] = rq
        dic_WQKSJD["开始"] = '15:20:00'
        dic_WQKSJD["结束"] = XWCDR_SJ
        if(XWCDR_SJ=='nan'):
            dic_WQKSJD["结束"] = '18:30:00'
            XWCDR_SJ='18:30:00'
        dic_WQKSJD["签到"] = ''
        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        ddjlData = dataDingding[(dataDingding['姓名'] == xm) & (dataDingding['日期'] == rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：", "15:20:00", "结束时间：", XWCDR_SJ)
            print(ddjlData)
            count_ddjl = ddjlData.shape[0]
            j = 0
            while (j < count_ddjl):
                ddjl = ddjlData.values[j]
                dksj = ddjl[4]   #李墡
                if (XWCDR_SJ > dksj and '15:20:00' < dksj):
                    print("打卡时间：", dksj, "在出去时间：", "15:20:00", "进来时间：", XWCDR_SJ, "之间")
                    dic_WQKSJD["签到"] = dksj
                    wqk = 1
                    dic_FaceDingding["下午迟到"] = "未迟到-钉钉"
                    dic_FaceDingding["下午迟到卡"] = dksj
                    break
                j = j + 1
        if (wqk == 0):
            dic_FaceDingding["下午迟到"] = "迟到-钉钉"

        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        print("筛选下午迟到事假数据")
        print(shijiajlData)
        print("筛选下午迟到事假数据")
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            t = 0
            while (t < count_shijiajl):
                shijiajl = shijiajlData.values[t]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                print("事假开始时间：", startTime, "结束时间：", endTime, "下午迟到入时间：", XWCDR_SJ)
                if (startTime <= '15:30:00' and endTime >= XWCDR_SJ):
                    shijia = 1
                    dic_FaceDingding["下午迟到"] = "未迟到-事假"
                    print("下午未迟到-事假")
                    break
                else:
                    pass
                t = t + 1
        list_XYWQKSJD.append(dic_WQKSJD)
    print('---------------------- 结束处理下午迟到待确认数据，与钉钉匹配签到数据和事假数据-------------------------')


    print('\n---------------------- 开始处理下午旷工数据，与钉钉匹配签到数据,匹配事假数据-------------------------')
    if (XWKG =='旷工' or XWKG=='待确认'):
        dic_WQKSJD = {}
        dic_WQKSJD["姓名"] = xm
        dic_WQKSJD["日期"] = rq
        dic_WQKSJD["开始"] = '15:30:00'
        dic_WQKSJD["结束"] = '18:30:00'
        dic_WQKSJD["签到"] = ''
        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        ddjlData = dataDingding[(dataDingding['姓名'] == xm) & (dataDingding['日期'] == rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：", "15:30:00", "结束时间：", "18:30:00")
            print(ddjlData)
            count_ddjl = ddjlData.shape[0]
            j = 0
            while (j < count_ddjl):
                ddjl = ddjlData.values[j]
                dksj = ddjl[4]
                #if ('18:30:00' >= dksj and '15:30:00' <= dksj):
                if ('18:30:00' >= dksj and '15:25:00' <= dksj):
                    print("打卡时间：", dksj, "在出去时间：", "15:30:00", "进来时间：", "18:30:00", "之间")
                    dic_WQKSJD["签到"] = dksj
                    wqk = 1
                    dic_FaceDingding["下午旷工"] = "未旷工-钉钉"
                    dic_FaceDingding["下午旷工卡"] = dksj
                    break
                j = j + 1
        if (wqk == 0):
            dic_FaceDingding["下午旷工"] = "旷工-钉钉"
            dic_FaceDingding["下午迟到"] = ''
            dic_FaceDingding["下午早退"] = ''
            dic_FaceDingding["下午早退出"]=''
        list_XYWQKSJD.append(dic_WQKSJD)

        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            k = 0
            while (k < count_shijiajl):
                shijiajl = shijiajlData.values[k]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                if (startTime <= '15:30:00' and endTime == '18:30:00'):
                    shijia = 1
                    dic_FaceDingding["下午旷工"] = "未旷工-事假"
                    break
                k = k + 1
    print('---------------------- 结束处理下午旷工数据，与钉钉匹配签到数据,匹配事假数据-------------------------')


    print('\n---------------------- 开始处理上午旷工数据，与钉钉匹配签到数据,匹配事假数据-------------------------')
    if (SWKG == '旷工'):
        shijia = 0  # 事假情况，符合要求为1，不符合要求为0，初始化为0
        shijiajlData = dataShijia[(dataShijia['姓名'] == xm) & (dataShijia['日期'] == rq)]
        if (shijiajlData.empty == False):
            count_shijiajl = shijiajlData.shape[0]
            k = 0
            while (k < count_shijiajl):
                shijiajl = shijiajlData.values[k]
                startTime = shijiajl[2]
                endTime = shijiajl[3]
                if (startTime == '09:30:00' and endTime >= '13:00:00'):
                    shijia = 1
                    dic_FaceDingding["上午旷工"] = "未旷工-事假"
                    break
                    k = k + 1

        wqk = 0  # 外勤卡情况，符合要求为1，不符合要求为0，初始化为0
        ddjlData = dataDingding[(dataDingding['姓名'] == xm) & (dataDingding['日期'] == rq)]
        if (ddjlData.empty == False):
            print("姓名：", xm, "日期：", rq, "开始时间：", "07:00:00", "结束时间：", "13:00:00")
            print(ddjlData)
            count_ddjl = ddjlData.shape[0]
            j = 0
            while (j < count_ddjl):
                ddjl = ddjlData.values[j]
                dksj = ddjl[4]
                if ('13:00:00' > dksj and '07:00:00' < dksj):
                    print("打卡时间：", dksj, "在出去时间：", "07:00:00", "进来时间：", "13:00:00", "之间")
                    #dic_WQKSJD["签到"] = dksj   #李墡注释掉了
                    wqk = 1
                    dic_FaceDingding["上午旷工"] = "未旷工-钉钉"
                    dic_FaceDingding["上午旷工卡"] = dksj
                    dic_FaceDingding["上午迟到"] = "迟到-钉钉"
                    break
                j = j + 1
        if (wqk == 0):
            dic_FaceDingding["上午旷工"] = "旷工-钉钉"
            dic_FaceDingding["上午迟到"] = ''
            dic_FaceDingding["上午早退"] = ''
            dic_FaceDingding["上午早退出"]=''
    print('---------------------- 结束处理上午旷工数据，与钉钉匹配签到数据,匹配事假数据-------------------------')

    list_FaceDingding.append(dic_FaceDingding)
    i=i+1

data_XYWQSJD=pd.DataFrame(list_XYWQKSJD)
print(data_XYWQSJD)
writer = pd.ExcelWriter(lj_zzjglj+'/需要外勤卡时间段.xlsx')
data_XYWQSJD.to_excel(writer,index=False)
writer.save()
writer.close()

data_FaceDingding=pd.DataFrame(list_FaceDingding)
print(data_FaceDingding)
writer = pd.ExcelWriter(lj_zzjglj+'/人脸钉钉统计数据.xlsx')
data_FaceDingding.to_excel(writer,index=False)
writer.save()
writer.close()

data_OvertimeDingding=pd.DataFrame(list_OvertimeDingding)
print(data_OvertimeDingding)
writer = pd.ExcelWriter(lj_zzjglj+'/出入超时钉钉未签到统计数据.xlsx')
data_OvertimeDingding.to_excel(writer,index=False)
writer.save()
writer.close()

#早到数据
data_ZaoDao=data_FaceDingding[(data_FaceDingding['早到']=='早到') | (data_FaceDingding['早到']=='早到-钉钉')][['姓名','日期']]
print('--------------------早到名单-----------------------')
print(data_ZaoDao)
data_ZaoDaoSums=data_ZaoDao.姓名.value_counts()
data_ZaoDaoSums=data_ZaoDaoSums.reset_index()
data_ZaoDaoSums.columns=['姓名','次数']
writer = pd.ExcelWriter(lj_zzjglj+'/早到情况.xlsx')
data_ZaoDao.to_excel(writer,sheet_name='早到名单',index=False)
data_ZaoDaoSums.to_excel(writer,sheet_name='早到次数统计',index=False)
writer.save()
writer.close()

#考勤异常名单
data_KaoQinYiChang=data_FaceDingding[(data_FaceDingding['上午旷工']=='旷工-钉钉') | (data_FaceDingding['下午旷工']=='旷工-钉钉') |
                              (data_FaceDingding['上午迟到']=='迟到')|  (data_FaceDingding['上午迟到']=='迟到-钉钉') |
                              (data_FaceDingding['上午早退']=='早退') | (data_FaceDingding['上午早退']=='早退-钉钉') |
                              (data_FaceDingding['下午迟到']=='迟到-钉钉') | (data_FaceDingding['下午迟到']=='迟到') |
                              (data_FaceDingding['下午早退']=='早退') | (data_FaceDingding['下午早退']=='早退-钉钉')][['姓名','日期',
                                                                                                       '上午旷工','下午旷工',
                                                                                                         '上午迟到',
                                                                                                         '上午早退',
                                                                                                         '下午迟到','下午早退']]
print(data_KaoQinYiChang)
writer = pd.ExcelWriter(lj_zzjglj+'/考勤异常名单.xlsx')
data_KaoQinYiChang.to_excel(writer,index=False)
writer.save()
writer.close()

print('---------读取考勤异常名单表格，对异常数据单元格进行颜色标准-----------')
wb= load_workbook(filename=lj_zzjglj+'/考勤异常名单.xlsx')
work=wb[wb.sheetnames[0]]
fill=PatternFill('solid',fgColor='FF0000')#填充红色
count_row = data_KaoQinYiChang.shape[0]
print('一共',count_row,'条考勤异常数据')
i=0
print('---准备遍历考勤异常数据---')
while(i<count_row):
    record=data_KaoQinYiChang.values[i]
    print('第',i,'条数据',record)
    SWKG=record[2]
    XWKG=record[3]
    SWCD=record[4]
    SWZT=record[5]
    XWCD=record[6]
    XWZT=record[7]
    if(SWKG=='旷工'):
        unit=work['C'+str(i+2)]
        unit.fill=fill
    if(XWKG=='旷工-钉钉'):
        unit=work['D'+str(i+2)]
        unit.fill=fill
    if(SWCD=='迟到' or SWCD=='迟到-钉钉'):
        unit=work['E'+str(i+2)]
        unit.fill=fill
    if(SWZT=='早退-钉钉'):
        unit=work['F'+str(i+2)]
        unit.fill=fill
    if(XWCD=='迟到-钉钉'):
        unit=work['G'+str(i+2)]
        unit.fill=fill
    if(XWZT=='早退-钉钉'):
        unit=work['H'+str(i+2)]
        unit.fill=fill
    i=i+1
wb.close()
wb.save(lj_zzjglj+'/考勤异常名单(标红).xlsx')
